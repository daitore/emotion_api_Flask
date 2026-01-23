from flask import Flask, request, jsonify, send_file, render_template
from io import BytesIO
from openpyxl import Workbook
from emotion import judge_emotion  # ← ロジックを別ファイルから読み込み

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False


# 画面
@app.get("/")
def index():
    return render_template("index.html")


# API（JSON）
@app.post("/api/emotion")
def api_emotion():
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text が空です"}), 400

    emotion, scores = judge_emotion(text)
    # 返すのは emotion だけ（シンプル）
    # 例：emotion が "主：喜び / 副：不安" みたいに返ってくる想定
    label = emotion

    main = "中立"
    sub = None

    if emotion.startswith("主："):
        main = emotion.split("主：")[1].split(" / ")[0].strip()
        if "副：" in emotion:
            sub = emotion.split("副：")[1].strip()
    else:
        main = emotion

    return jsonify({
        "label": label,  # 表示用
        "main": main,  # CSS用（主）
        "sub": sub,  # 副（使いたい時用）
        "scores": scores
    })


# Excelダウンロード
@app.get("/download.xlsx")
def download_xlsx():
    text = (request.args.get("text") or "").strip()
    emotion, scores = judge_emotion(text)

    wb = Workbook()
    ws = wb.active
    ws.title = "result"
    ws["A1"] = "text"
    ws["B1"] = text
    ws["A2"] = "emotion"
    ws["B2"] = emotion

    ws["A4"] = "scores"
    row = 5
    for k, v in scores.items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="emotion_result.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
